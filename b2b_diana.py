import openpyxl
import PySimpleGUI as sg
import os

<<<<<<< HEAD
=======
# pyinstaller --onefile -w "b2b_diana.py" --windowed

>>>>>>> caab72a (reacomodo de git local)
'''
> Obtener UUID por cada proveedor
> Contar cuantos proveedores
'''

# obtener la ubicación y nomre del archivo excel
path = os.path.normpath(sg.popup_get_file("Selecciona el archivo excel a leer:"))

# Insancias de objetos
wb_obj = openpyxl.load_workbook(path, data_only=True)
sheet_obj = wb_obj["Reporte Facturas"]


# Obtener encabezados e indice de la columna "Nombre del proveedor" y "UUID"
encabezados = []
for i in range(1, sheet_obj.max_column + 1):
	cell_obj = sheet_obj.cell(row=1, column=i)
	encabezados.append(cell_obj.value)
index_proveedor = encabezados.index("Nombre del proveedor")+1
index_UUID = encabezados.index("UUID")+1

# Obtener lista de proveedores y remover duplicados
proveedores = []
for i in range(1, sheet_obj.max_row +1):
	if i == 1:
		continue
	cell_obj = sheet_obj.cell(row=i, column=index_proveedor)
	proveedores.append(cell_obj.value)
proveedores_filt = sorted(list(set(proveedores)))
proveedores_comp = sorted(proveedores)


#Obtener UUID =================================================
UUID = {}

# Rellenar diccionario conproveedores a listas vacías
dict_prov_filt = {}
for prov in proveedores_filt:
	if prov not in dict_prov_filt:
		dict_prov_filt[prov] = 1
		UUID[prov] = []
	else:
		dict_prov_filt[prov] += 1

for prov in proveedores_comp:
	if prov not in dict_prov_filt:
		dict_prov_filt[prov] = []

# REllenar listas con UUIDs
for i in range(1, sheet_obj.max_row+1):
	if i == 1:
		continue
	UUID_obj = sheet_obj.cell(row=i, column=index_UUID)
	prov_obj = sheet_obj.cell(row=i, column=index_proveedor)
	UUID[prov_obj.value].append(UUID_obj.value)

# Convertir lista de UUIDs en STR
for prov in UUID:
	UUID[prov] = ",".join(UUID[prov])


# Modificar la segunda hoja para mostrar los resultados

ws_python = wb_obj["Hoja1"]

for i in range(1, ws_python.max_row+1):
	if i == 0: continue
	else:
		encabezados_out = []
		for i in range(1, ws_python.max_column + 1):
			cell_obj = ws_python.cell(row=1, column=i)
			encabezados_out.append(cell_obj.value)
		index_proveedor_out = encabezados_out.index("Razon social")+1
		index_UUID_out = encabezados_out.index("UUID")+1

# Escribir y guardar los datos
for i, prov in enumerate(UUID):
	cell_obj = ws_python.cell(row = i+2, column=1)
	cell_obj.value = prov

	cell_obj = ws_python.cell(row = i+2, column=3)
	cell_obj.value = UUID[prov]
wb_obj.save(path)

sg.popup_ok(f"Se han guardado los cambios al archivo:\n{path}")